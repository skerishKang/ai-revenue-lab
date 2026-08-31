from __future__ import annotations

import base64
import binascii
from io import BytesIO
from pathlib import PurePath
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader

from .documents import DocumentAttachment, MAX_DOCUMENT_CHARS, _safe_name, _validate_text
from .ooxml_stdlib import (
    MAX_ARCHIVE_BYTES,
    OOXMLExtractionError,
    extract_docx_text,
    extract_pptx_text,
    validate_ooxml_archive,
)

MAX_BINARY_DOCUMENT_BYTES = MAX_ARCHIVE_BYTES
MAX_PDF_PAGES = 80
MAX_XLSX_SHEETS = 20
MAX_XLSX_ROWS = 500
MAX_XLSX_COLUMNS = 50
MAX_XLSX_NONEMPTY_CELLS = 5_000

BINARY_DOCUMENT_MEDIA = {
    "application/pdf": {".pdf"},
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": {".docx"},
    "application/vnd.openxmlformats-officedocument.presentationml.presentation": {".pptx"},
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {".xlsx"},
}


class BinaryDocumentValidationError(ValueError):
    pass


def _validate_name_and_media(name_value: Any, media_type: Any) -> tuple[str, str]:
    try:
        name = _safe_name(name_value)
    except Exception as exc:
        raise BinaryDocumentValidationError("문서 파일 이름 형식이 올바르지 않습니다.") from exc
    if not isinstance(media_type, str) or media_type not in BINARY_DOCUMENT_MEDIA:
        raise BinaryDocumentValidationError("PDF, DOCX, PPTX, XLSX 문서만 이 형식으로 첨부할 수 있습니다.")
    suffix = PurePath(name.lower()).suffix
    if suffix not in BINARY_DOCUMENT_MEDIA[media_type]:
        raise BinaryDocumentValidationError("문서 확장자와 파일 형식이 일치하지 않습니다.")
    return name, media_type


def _decode_payload(payload: Any) -> bytes:
    if not isinstance(payload, str) or not payload:
        raise BinaryDocumentValidationError("문서 데이터가 올바르지 않습니다.")
    if len(payload) > ((MAX_BINARY_DOCUMENT_BYTES + 2) // 3) * 4:
        raise BinaryDocumentValidationError("PDF·Office 문서는 2 MiB 이하만 첨부할 수 있습니다.")
    try:
        decoded = base64.b64decode(payload, validate=True)
    except (binascii.Error, ValueError) as exc:
        raise BinaryDocumentValidationError("문서 데이터가 올바르지 않습니다.") from exc
    if not decoded:
        raise BinaryDocumentValidationError("빈 문서는 첨부할 수 없습니다.")
    if len(decoded) > MAX_BINARY_DOCUMENT_BYTES:
        raise BinaryDocumentValidationError("PDF·Office 문서는 2 MiB 이하만 첨부할 수 있습니다.")
    return decoded


def _bounded_append(parts: list[str], value: str) -> None:
    if not value:
        return
    parts.append(value)
    if sum(len(part) for part in parts) + max(0, len(parts) - 1) > MAX_DOCUMENT_CHARS:
        raise BinaryDocumentValidationError("문서에서 추출한 텍스트가 40,000자를 초과합니다.")


def _extract_pdf_text(payload: bytes) -> str:
    if not payload.startswith(b"%PDF-"):
        raise BinaryDocumentValidationError("PDF 형식과 실제 파일 내용이 일치하지 않습니다.")
    try:
        reader = PdfReader(BytesIO(payload), strict=True)
        if reader.is_encrypted:
            raise BinaryDocumentValidationError("암호화된 PDF는 현재 첨부할 수 없습니다.")
        page_count = len(reader.pages)
        if page_count < 1:
            raise BinaryDocumentValidationError("PDF에서 읽을 페이지를 찾지 못했습니다.")
        if page_count > MAX_PDF_PAGES:
            raise BinaryDocumentValidationError("PDF는 80페이지 이하만 첨부할 수 있습니다.")
        parts: list[str] = []
        for page in reader.pages:
            extracted = page.extract_text() or ""
            if extracted.strip():
                _bounded_append(parts, extracted.strip())
        text = "\n\n".join(parts)
        if not text.strip():
            raise BinaryDocumentValidationError("이 PDF에서는 텍스트를 읽지 못했습니다. 스캔 문서 OCR은 아직 지원하지 않습니다.")
        return text
    except BinaryDocumentValidationError:
        raise
    except Exception as exc:
        raise BinaryDocumentValidationError("PDF를 안전하게 읽지 못했습니다.") from exc


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def _extract_xlsx_text(payload: bytes) -> str:
    try:
        validate_ooxml_archive(payload)
        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True, keep_links=False)
    except OOXMLExtractionError as exc:
        raise BinaryDocumentValidationError("XLSX 압축 구조를 안전하게 읽지 못했습니다.") from exc
    except Exception as exc:
        raise BinaryDocumentValidationError("XLSX 파일을 안전하게 읽지 못했습니다.") from exc

    try:
        if len(workbook.sheetnames) < 1:
            raise BinaryDocumentValidationError("XLSX에서 워크시트를 찾지 못했습니다.")
        if len(workbook.sheetnames) > MAX_XLSX_SHEETS:
            raise BinaryDocumentValidationError("XLSX는 워크시트 20개 이하만 첨부할 수 있습니다.")

        parts: list[str] = []
        nonempty_cells = 0
        for worksheet in workbook.worksheets:
            max_row = int(worksheet.max_row or 0)
            max_column = int(worksheet.max_column or 0)
            if max_row > MAX_XLSX_ROWS or max_column > MAX_XLSX_COLUMNS:
                raise BinaryDocumentValidationError("XLSX는 시트당 500행 × 50열 범위까지만 읽을 수 있습니다.")
            for row in worksheet.iter_rows(max_row=min(max_row, MAX_XLSX_ROWS), max_col=min(max_column, MAX_XLSX_COLUMNS)):
                for cell in row:
                    value = _cell_text(cell.value)
                    if not value:
                        continue
                    nonempty_cells += 1
                    if nonempty_cells > MAX_XLSX_NONEMPTY_CELLS:
                        raise BinaryDocumentValidationError("XLSX의 값이 너무 많습니다.")
                    _bounded_append(parts, f"[{worksheet.title}!{cell.coordinate}] {value}")
        text = "\n".join(parts)
        if not text.strip():
            raise BinaryDocumentValidationError("XLSX에서 읽을 값을 찾지 못했습니다.")
        return text
    finally:
        workbook.close()


def _extract_text(media_type: str, payload: bytes) -> str:
    if media_type == "application/pdf":
        return _extract_pdf_text(payload)
    if media_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        try:
            text = extract_docx_text(payload)
        except OOXMLExtractionError as exc:
            raise BinaryDocumentValidationError("DOCX 파일을 안전하게 읽지 못했습니다.") from exc
        if not text.strip():
            raise BinaryDocumentValidationError("DOCX에서 읽을 텍스트를 찾지 못했습니다.")
        return text
    if media_type == "application/vnd.openxmlformats-officedocument.presentationml.presentation":
        try:
            text = extract_pptx_text(payload)
        except OOXMLExtractionError as exc:
            raise BinaryDocumentValidationError("PPTX 파일을 안전하게 읽지 못했습니다.") from exc
        if not text.strip():
            raise BinaryDocumentValidationError("PPTX에서 읽을 텍스트를 찾지 못했습니다.")
        return text
    if media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return _extract_xlsx_text(payload)
    raise BinaryDocumentValidationError("지원하지 않는 문서 형식입니다.")


def parse_binary_document_item(item: Any) -> DocumentAttachment:
    if not isinstance(item, dict) or set(item) != {"type", "name", "media_type", "base64"}:
        raise BinaryDocumentValidationError("지원하지 않는 바이너리 문서 첨부 항목이 있습니다.")
    if item.get("type") != "document":
        raise BinaryDocumentValidationError("문서 첨부 형식이 올바르지 않습니다.")
    name, media_type = _validate_name_and_media(item.get("name"), item.get("media_type"))
    payload = _decode_payload(item.get("base64"))
    text = _extract_text(media_type, payload)
    try:
        normalized = _validate_text(text)
    except Exception as exc:
        raise BinaryDocumentValidationError("문서에서 추출한 텍스트를 사용할 수 없습니다.") from exc
    return DocumentAttachment(name=name, media_type=media_type, text=normalized, byte_size=len(payload))
