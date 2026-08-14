#!/usr/bin/env python3
"""Generate the Business 35 pilot quote template XLSX.

Sheets: Instructions, Customer Scope, Offer A, Offer B, Offer C,
Optional Items, Assumptions, Approval.

Calculates 공급가액 / VAT / 합계 / 지급 단계. Warns when a quoted amount is
outside the standard price range.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "docs/commercial/business-35-ai-media-education-dx/customer-package/Business35_Pilot_Quote_Template.xlsx"

NAVY = "1F3A5F"
BLUE = "2E5E8C"
LIGHT = "F2F4F7"
ACCENT = "C27B2D"
RED = "B02A2A"
WHITE = "FFFFFF"

THIN = Side(style="thin", color="B0B7C0")


def set_cell(ws, row, col, value, bold=False, size=11, color=None, fill=None, align=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Malgun Gothic", size=size, bold=bold, color=color or "1F3A5F")
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    c.alignment = Alignment(horizontal=align or "left", vertical="center", wrap_text=wrap)
    return c


def build():
    wb = Workbook()

    # ---- Instructions ----
    ws = wb.active
    ws.title = "Instructions"
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 90
    rows = [
        ("Business 35 · 견적 검토 템플릿", ""),
        ("공급자: 파디엠", ""),
        ("제공 및 계약 주체: 파디엠", ""),
        ("견적 발행 주체: 파디엠", ""),
        ("", ""),
        ("사용 방법", "각 Offer sheet에 금액(원, VAT 별도)과 인원·기간을 입력합니다."),
        ("계산", "공급가액 / VAT(10%) / 합계가 자동 계산됩니다. 지급 단계(착수금 50%·잔금 50%)를 확인합니다."),
        ("경고", "표준 가격 범위를 벗어나는 금액을 입력하면 해당 셀에 경고가 표시됩니다."),
        ("표준 가격", "A 초기형 300만–500만원 · A 확장형 500만–800만원\nB1 1,000만–1,500만원 · B2 1,500만–2,500만원 · C 월 300만–600만원"),
        ("상태", "견적 검토 템플릿 — 계약 확정 문서가 아닙니다."),
        ("가격 정책", "시장 검증 전 자사 가격 가설 · 범위와 인원·기간에 따라 최종 견적 · VAT 조건은 최종 견적서에서 확정"),
        ("", ""),
        ("주의", "발송 전 공식 사업자 정보 입력 필요 (사업자등록번호·대표자·주소·연락처·계좌)."),
        ("주의", "본 템플릿은 고객에게 발송 전 최종 승인과 전문 법률·계약 검토가 필요합니다."),
    ]
    for i, (a, b) in enumerate(rows, start=1):
        set_cell(ws, i, 1, a, bold=i in (1,), fill=LIGHT if i > 4 else None)
        set_cell(ws, i, 2, b, size=10, color="333333" if b else None, fill=LIGHT if i > 4 else None, wrap=True)
    # B9 (표준 가격): two lines, tall enough to show all prices
    ws.row_dimensions[9].height = 34

    # ---- Customer Scope ----
    ws = wb.create_sheet("Customer Scope")
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 60
    fields = [
        ("고객 조직명", "[고객 조직명]"),
        ("파일럿 팀", "[팀명·인원]"),
        ("대상 업무 1건", "[업무명]"),
        ("파일럿 기간", "6주"),
        ("참여 인원", "[인원]"),
        ("승인 책임자", "[직책]"),
        ("운영 책임자", "[직책]"),
        ("시작 예정일", "[YYYY-MM-DD]"),
        ("종료 예정일", "[YYYY-MM-DD]"),
    ]
    set_cell(ws, 1, 1, "Customer Scope", bold=True, size=14, fill=LIGHT)
    for i, (a, b) in enumerate(fields, start=3):
        set_cell(ws, i, 1, a, bold=True, fill=LIGHT)
        set_cell(ws, i, 2, b, color="333333")

    # ---- Offer sheets ----
    offers = {
        "Offer A": ("A · AI 업무전환 진단 워크숍", 3_000_000, 8_000_000, "초기형 300만–500만원 / 확장형 500만–800만원"),
        "Offer B1": ("B1 · 6주 디자인 파트너 파일럿", 10_000_000, 15_000_000, "1,000만–1,500만원"),
        "Offer B2": ("B2 · 6주 표준 파일럿", 15_000_000, 25_000_000, "1,500만–2,500만원"),
        "Offer C": ("C · 조직 운영 자문 (월)", 3_000_000, 6_000_000, "월 300만–600만원"),
    }
    for sheet, (title, lo, hi, range_desc) in offers.items():
        ws = wb.create_sheet(sheet)
        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 16
        set_cell(ws, 1, 1, title, bold=True, size=14, fill=LIGHT)
        set_cell(ws, 2, 1, range_desc, size=10, color=ACCENT)
        set_cell(ws, 3, 1, "항목", bold=True, fill=NAVY, color=WHITE)
        set_cell(ws, 3, 2, "금액(원, VAT 별도)", bold=True, fill=NAVY, color=WHITE)
        set_cell(ws, 3, 3, "인원", bold=True, fill=NAVY, color=WHITE)
        set_cell(ws, 3, 4, "기간", bold=True, fill=NAVY, color=WHITE)
        set_cell(ws, 4, 1, "기본 견적", bold=True)
        set_cell(ws, 4, 2, (lo + hi) // 2)
        set_cell(ws, 4, 3, "")
        set_cell(ws, 4, 4, "")
        set_cell(ws, 5, 1, "공급가액", bold=True, fill=LIGHT)
        set_cell(ws, 5, 2, f"=B4")
        set_cell(ws, 6, 1, "VAT (10%)", fill=LIGHT)
        set_cell(ws, 6, 2, "=ROUND(B5*0.1,0)")
        set_cell(ws, 7, 1, "합계 (VAT 포함)", bold=True, fill=LIGHT)
        set_cell(ws, 7, 2, "=B5+B6")
        set_cell(ws, 9, 1, "지급 단계", bold=True, fill=NAVY, color=WHITE)
        set_cell(ws, 10, 1, "착수금 50%", fill=LIGHT)
        set_cell(ws, 10, 2, "=ROUND(B5*0.5,0)")
        set_cell(ws, 11, 1, "최종 산출물 검수 후 잔금 50%", fill=LIGHT)
        set_cell(ws, 11, 2, "=ROUND(B5*0.5,0)")
        set_cell(ws, 13, 1, "참고: 견적 검토 템플릿 — 계약 확정 문서 아님", size=9, color="555A60")

        # Range warning via conditional formatting
        from openpyxl.formatting.rule import FormulaRule
        ws.conditional_formatting.add(
            "B4:B4",
            FormulaRule(
                formula=[f"OR(B4<{lo},B4>{hi})"],
                fill=PatternFill("solid", fgColor=RED),
                font=Font(color=WHITE, bold=True),
                stopIfTrue=True,
            ),
        )
        # Note text under the amount
        set_cell(ws, 15, 1, f"표준 범위: {lo:,} – {hi:,} 원 (VAT 별도). 범위 밖 값은 빨간 경고 표시.", size=9, color="555A60")

    # ---- Optional Items ----
    ws = wb.create_sheet("Optional Items")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    set_cell(ws, 1, 1, "Optional Items", bold=True, size=14, fill=LIGHT)
    set_cell(ws, 2, 1, "항목", bold=True, fill=NAVY, color=WHITE)
    set_cell(ws, 2, 2, "금액(원)", bold=True, fill=NAVY, color=WHITE)
    set_cell(ws, 2, 3, "선택", bold=True, fill=NAVY, color=WHITE)
    set_cell(ws, 3, 1, "추가 팀 확장 워크숍")
    set_cell(ws, 3, 3, "☐")
    set_cell(ws, 4, 1, "월간 운영 자문 (추가 개월)")
    set_cell(ws, 4, 3, "☐")
    set_cell(ws, 5, 1, "맞춤 교육 자료 제작")
    set_cell(ws, 5, 3, "☐")
    set_cell(ws, 7, 1, "선택 항목 금액은 범위 확인 후 별도 승인 필요", size=9, color="555A60")

    # ---- Assumptions ----
    ws = wb.create_sheet("Assumptions")
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 70
    items = [
        ("범위", "1팀·1핵심 업무 · 6주 파일럿 (Week 0–6)"),
        ("참여", "팀 주당 2–4시간 · 운영 책임자 1인 지정"),
        ("산출물", "기준선 진단 · 교육 · 워크플로 재설계 · 성과·위험 보고 · 운영 플레이북"),
        ("제외", "전사 확산 · 시스템 구축 · AI 도구 라이선스 구매 · 법률 자문"),
        ("데이터", "개인정보 원문은 외부 AI 입력 금지 · 입력 금지 자료 준수"),
        ("가격", "시장 검증 전 자사 가격 가설 · 실제 견적은 범위 확인 후 별도 승인"),
        ("VAT", "VAT 조건은 최종 견적서에서 확정"),
    ]
    set_cell(ws, 1, 1, "Assumptions", bold=True, size=14, fill=LIGHT)
    for i, (a, b) in enumerate(items, start=3):
        set_cell(ws, i, 1, a, bold=True, fill=LIGHT)
        set_cell(ws, i, 2, b, color="333333", wrap=True)

    # ---- Approval ----
    ws = wb.create_sheet("Approval")
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 45
    set_cell(ws, 1, 1, "Approval", bold=True, size=14, fill=LIGHT)
    fields = [
        ("제공 및 계약 주체", "파디엠"),
        ("공급자", "파디엠"),
        ("사업자등록번호", "발송 전 공식 사업자 정보 입력 필요"),
        ("대표자·주소·연락처", "발송 전 공식 사업자 정보 입력 필요"),
        ("입금 계좌 (은행명·예금주)", "발송 전 공식 사업자 정보 입력 필요"),
        ("견적 준비자", "[이름/직책]"),
        ("견적 승인자 (제공자)", "[이름/직책]"),
        ("고객 승인자", "[직책]"),
        ("승인일", "[YYYY-MM-DD]"),
        ("상태", "견적 검토 중 (계약 확정 아님)"),
        ("법률·계약 검토", "전문 법률·계약 검토 필요 — 최종 발송 전 확인"),
    ]
    for i, (a, b) in enumerate(fields, start=3):
        set_cell(ws, i, 1, a, bold=True, fill=LIGHT)
        set_cell(ws, i, 2, b, color="333333", wrap=True)
    # B13 (법률·계약 검토): full Korean phrase visible, no right clipping
    ws.row_dimensions[13].height = 32
    note = set_cell(ws, 15, 1, "안내: 사업자등록번호·대표자·주소·연락처·계좌 정보는 발송 전 공식 확인 후 입력합니다.", size=9, color="555A60")
    ws.merge_cells("A15:B15")

    wb.save(OUT)
    print(f"saved {OUT}")


if __name__ == "__main__":
    build()
