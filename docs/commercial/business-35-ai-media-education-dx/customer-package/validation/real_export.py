#!/usr/bin/env python3
"""Real document exporters for B35 Lane B final artifacts (render-fidelity).

Final customer PDFs must be produced by a REAL document renderer/exporter:

- PPTX -> PDF: Microsoft PowerPoint COM (preferred) or LibreOffice Impress headless
- DOCX -> PDF: Microsoft Word COM (preferred) or LibreOffice Writer headless
- XLSX sheets -> PDF: Microsoft Excel COM (preferred) or LibreOffice Calc headless

Independently recomposed PDFs (fpdf2/reportlab-style fallback) are FORBIDDEN.
When no real exporter is available every entry point raises
RealExportUnavailable and the caller must fail closed with
REAL_DOCUMENT_EXPORT=UNAVAILABLE_BLOCKING or REAL_XLSX_RENDER=UNAVAILABLE_BLOCKING.

Exported PDFs are normalized to deterministic bytes (fixed metadata dates,
document IDs, and XMP packet stamps) so OUTPUT_HASHES remain repeatable.
"""

from __future__ import annotations

import json
import re
import shutil
import subprocess
import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

ROOT = Path(__file__).resolve().parent.parent
REAL_EXPORT_DIR = Path(__file__).resolve().parent / ".real_export"
PROVENANCE_PATH = REAL_EXPORT_DIR / "provenance.json"

FIXED_PDF_DATE = "D:20260903000000+00'00'"
FIXED_XMP_DATE = "2026-09-03T00:00:00+00:00"
FIXED_UUID = "00000000-0000-0000-0000-000000000000"
FIXED_ID_HEX = "0" * 32


class RealExportUnavailable(RuntimeError):
    """Raised when no real document exporter exists on this machine."""


# ---------------------------------------------------------------- provenance

def read_provenance() -> dict:
    if PROVENANCE_PATH.is_file():
        try:
            return json.loads(PROVENANCE_PATH.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def write_provenance(**fields) -> dict:
    REAL_EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    prov = read_provenance()
    prov.update(fields)
    PROVENANCE_PATH.write_text(json.dumps(prov, ensure_ascii=False, indent=2, sort_keys=True),
                               encoding="utf-8")
    return prov


# ------------------------------------------------------- deterministic scrub

def normalize_pdf_determinism(path: Path | str) -> None:
    """Pin PDF metadata dates, trailer IDs, and XMP stamps to fixed values."""
    try:
        import fitz
    except Exception as e:
        raise RealExportUnavailable(f"PyMuPDF missing for PDF normalization: {e}")
    path = str(path)
    doc = fitz.open(path)
    md = dict(doc.metadata or {})
    md["creationDate"] = FIXED_PDF_DATE
    md["modDate"] = FIXED_PDF_DATE
    doc.set_metadata(md)
    tmp = path + ".norm"
    doc.save(tmp, garbage=4, deflate=True)
    doc.close()
    raw = Path(tmp).read_bytes()
    raw = re.sub(rb"/CreationDate\s*\([^)]*\)",
                 b"/CreationDate (" + FIXED_PDF_DATE.encode() + b")", raw)
    raw = re.sub(rb"/ModDate\s*\([^)]*\)",
                 b"/ModDate (" + FIXED_PDF_DATE.encode() + b")", raw)
    raw = re.sub(rb"/ID\s*\[\s*<[0-9A-Fa-f]+>\s*<[0-9A-Fa-f]+>\s*\]",
                 (" /ID [<%s><%s>] " % (FIXED_ID_HEX, FIXED_ID_HEX)).encode(), raw)
    raw = re.sub(rb"<xmp:CreateDate>[^<]*</xmp:CreateDate>",
                 ("<xmp:CreateDate>%s</xmp:CreateDate>" % FIXED_XMP_DATE).encode(), raw)
    raw = re.sub(rb"<xmp:ModifyDate>[^<]*</xmp:ModifyDate>",
                 ("<xmp:ModifyDate>%s</xmp:ModifyDate>" % FIXED_XMP_DATE).encode(), raw)
    raw = re.sub(rb"<xmpMM:DocumentID>[^<]*</xmpMM:DocumentID>",
                 ("<xmpMM:DocumentID>uuid:%s</xmpMM:DocumentID>" % FIXED_UUID).encode(), raw)
    raw = re.sub(rb"<xmpMM:InstanceID>[^<]*</xmpMM:InstanceID>",
                 ("<xmpMM:InstanceID>uuid:%s</xmpMM:InstanceID>" % FIXED_UUID).encode(), raw)
    Path(path).write_bytes(raw)
    Path(tmp).unlink()
    print(f"normalized {Path(path).name} (fixed PDF/XMP metadata)")


def pdf_producer(path: Path | str) -> str:
    import fitz
    doc = fitz.open(str(path))
    prod = (doc.metadata or {}).get("producer", "") or ""
    doc.close()
    return prod


# ------------------------------------------------------------------- backends

def _libreoffice() -> str | None:
    return shutil.which("soffice") or shutil.which("libreoffice")


def _lo_convert(src: Path, out_dir: Path, expected: Path) -> bool:
    soffice = _libreoffice()
    if not soffice:
        return False
    try:
        result = subprocess.run(
            [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(out_dir), str(src)],
            capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=180,
        )
    except Exception as e:
        print(f"libreoffice failed for {src.name}: {e}")
        return False
    return result.returncode == 0 and expected.exists()


def _win32com():
    try:
        import win32com.client  # noqa: F401
        import win32com.client as wc
        return wc
    except Exception:
        return None


def export_pptx_to_pdf(src: Path, dst: Path) -> str:
    """Export an editable PPTX to PDF via a real renderer. Returns exporter name."""
    wc = _win32com()
    if wc is not None:
        app = None
        pres = None
        try:
            app = wc.DispatchEx("PowerPoint.Application")
            app.DisplayAlerts = 0
            pres = app.Presentations.Open(str(src), True, False, False)
            if dst.exists():
                dst.unlink()
            pres.SaveAs(str(dst), 32)  # ppSaveAsPDF
            pres.Close()
            pres = None
            app.Quit()
            app = None
        except Exception as e:
            try:
                if pres is not None:
                    pres.Close()
            except Exception:
                pass
            try:
                if app is not None:
                    app.Quit()
            except Exception:
                pass
            print(f"PowerPoint COM export failed for {src.name}: {e}")
        else:
            if dst.is_file():
                return "Microsoft PowerPoint COM"
    if _lo_convert(src, dst.parent, dst):
        return "LibreOffice Impress headless"
    raise RealExportUnavailable(
        f"no real PPTX exporter (PowerPoint COM / LibreOffice) for {src.name}")


def export_docx_to_pdf(src: Path, dst: Path) -> str:
    """Export a DOCX to PDF via a real renderer. Returns exporter name."""
    wc = _win32com()
    if wc is not None:
        app = None
        doc = None
        try:
            app = wc.DispatchEx("Word.Application")
            app.Visible = False
            app.DisplayAlerts = 0
            doc = app.Documents.Open(str(src), True)
            if dst.exists():
                dst.unlink()
            doc.SaveAs(str(dst), 17)  # wdFormatPDF
            doc.Close(False)
            doc = None
            app.Quit()
            app = None
        except Exception as e:
            try:
                if doc is not None:
                    doc.Close(False)
            except Exception:
                pass
            try:
                if app is not None:
                    app.Quit()
            except Exception:
                pass
            print(f"Word COM export failed for {src.name}: {e}")
        else:
            if dst.is_file():
                return "Microsoft Word COM"
    if _lo_convert(src, dst.parent, dst):
        return "LibreOffice Writer headless"
    raise RealExportUnavailable(
        f"no real DOCX exporter (Word COM / LibreOffice) for {src.name}")


def export_xlsx_sheets_pdf(src: Path, dst: Path) -> tuple[str, list[str], int]:
    """Export workbook sheets to PDF via a real spreadsheet renderer.

    In-memory FitToPages 1x1 is applied without saving back, so the source
    workbook bytes are never mutated. Returns (exporter, sheet_names, pdf_pages).
    """
    import fitz
    wc = _win32com()
    if wc is not None:
        app = None
        wb = None
        try:
            app = wc.DispatchEx("Excel.Application")
            app.Visible = False
            app.DisplayAlerts = False
            wb = app.Workbooks.Open(str(src), True)
            sheets = [ws.Name for ws in wb.Worksheets]
            for ws in wb.Worksheets:
                ws.PageSetup.FitToPagesWide = 1
                ws.PageSetup.FitToPagesTall = 1
                ws.PageSetup.Zoom = False
            if dst.exists():
                dst.unlink()
            wb.ExportAsFixedFormat(0, str(dst))  # xlTypePDF
            wb.Close(False)
            wb = None
            app.Quit()
            app = None
        except Exception as e:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if app is not None:
                    app.Quit()
            except Exception:
                pass
            print(f"Excel COM export failed for {src.name}: {e}")
        else:
            if dst.is_file():
                with fitz.open(str(dst)) as d:
                    pages = d.page_count
                if pages != len(sheets):
                    raise RealExportUnavailable(
                        f"Excel sheet/page mismatch for {src.name}: "
                        f"{len(sheets)} sheets vs {pages} pdf pages")
                return "Microsoft Excel COM", sheets, pages
    if _lo_convert(src, dst.parent, dst):
        with fitz.open(str(dst)) as d:
            pages = d.page_count
        # LibreOffice sheet order follows workbook order.
        from openpyxl import load_workbook
        wb = load_workbook(str(src), data_only=False, read_only=True)
        sheets = list(wb.sheetnames)
        wb.close()
        if pages != len(sheets):
            raise RealExportUnavailable(
                f"Calc sheet/page mismatch for {src.name}: "
                f"{len(sheets)} sheets vs {pages} pdf pages")
        return "LibreOffice Calc headless", sheets, pages
    raise RealExportUnavailable(
        f"no real XLSX renderer (Excel COM / LibreOffice) for {src.name}")
