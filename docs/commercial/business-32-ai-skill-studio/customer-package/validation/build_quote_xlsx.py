"""Build Business32_Pilot_Quote_Template.xlsx (8 sheets, formula-based)."""
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "Business32_Pilot_Quote_Template.xlsx",
)

BLUE = "2B6CB0"
ORANGE = "C05621"
YELLOW = "FFF3CD"
GRAY = "6B7280"

RANGES = {
    "A": (3000000, 5000000, "300만~500만원"),
    "B": (5000000, 8000000, "500만~800만원"),
    "C": (12000000, 20000000, "1,200만~2,000만원"),
}


def title(ws, text):
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=14, color=BLUE)


def header(ws, row, cols, fill=BLUE):
    for i, text in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=text)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=fill)


wb = Workbook()
ws_info = wb.active
ws_info.title = "안내"
title(ws_info, "Business 32 · Pilot Quote Template (안내)")
rows = [
    "이 견적 템플릿은 제안검토용입니다. 계약을 자동 확정하지 않습니다.",
    "모든 가격은 가설입니다. 확정 가격이 아닙니다.",
    "실제 고객·기관·견적서·내부 문서는 사용하지 않습니다. 합성 샘플 기반.",
    "사람 검토와 승인을 포함하는 서비스형 프론트엔드 파일럿입니다.",
    "시트: 안내 / 고객·업무 / Offer 선택 / 작업 범위 / 산출물 / 일정 / 견적 / 가정·제외사항",
]
for i, r in enumerate(rows, start=3):
    ws_info.cell(row=i, column=1, value=r)
ws_info.column_dimensions["A"].width = 90

ws_cust = wb.create_sheet("고객·업무")
title(ws_cust, "고객·업무")
header(ws_cust, 3, ["항목", "값(합성 예시)", "비고"])
cust_rows = [
    ("기관명", "가상 기관명 A", "실제 기관 사용 금지"),
    ("팀", "홍보팀", ""),
    ("선택 업무", "교육 프로그램 안내문 작성 및 검토", ""),
    ("실행자", "홍보 담당", ""),
    ("검토·승인자", "팀장 · 담당 책임자", ""),
    ("업무 빈도", "주 1회 (가상)", ""),
]
for i, (a, b, c) in enumerate(cust_rows, start=4):
    ws_cust.cell(row=i, column=1, value=a).font = Font(bold=True)
    ws_cust.cell(row=i, column=2, value=b)
    ws_cust.cell(row=i, column=3, value=c)
ws_cust.column_dimensions["A"].width = 18
ws_cust.column_dimensions["B"].width = 40
ws_cust.column_dimensions["C"].width = 30

ws_offer = wb.create_sheet("Offer 선택")
title(ws_offer, "Offer 선택")
header(ws_offer, 3, ["항목", "값", "가격 가설"])
ws_offer.cell(row=4, column=1, value="Offer").font = Font(bold=True)
ws_offer["B4"] = "B"
ws_offer.cell(row=4, column=3, value="=IF($B$4=\"A\",\"300만~500만원\",IF($B$4=\"C\",\"1,200만~2,000만원\",\"500만~800만원\"))")
dv = DataValidation(type="list", formula1='"A,B,C"', allow_blank=False)
dv.error = "A/B/C 중 하나를 선택하세요."
ws_offer.add_data_validation(dv)
dv.add(ws_offer["B4"])
ws_offer.column_dimensions["A"].width = 14
ws_offer.column_dimensions["B"].width = 10
ws_offer.column_dimensions["C"].width = 26

ws_scope = wb.create_sheet("작업 범위")
title(ws_scope, "작업 범위")
header(ws_scope, 3, ["범위 항목", "포함 여부", "비고"])
scope_rows = [
    ("업무 1개 선정", "Y", ""),
    ("입력자료·단계·증거 정의", "Y", ""),
    ("검토·승인 기준 정의", "Y", ""),
    ("합성 실행 (검증된 프론트엔드)", "Y", ""),
    ("스킬 카드 납품", "Y", ""),
    ("운영 가이드", "B/C 이상", ""),
]
for i, (a, b, c) in enumerate(scope_rows, start=4):
    ws_scope.cell(row=i, column=1, value=a)
    ws_scope.cell(row=i, column=2, value=b)
    ws_scope.cell(row=i, column=3, value=c)
ws_scope.column_dimensions["A"].width = 40
ws_scope.column_dimensions["B"].width = 12
ws_scope.column_dimensions["C"].width = 30

ws_deliv = wb.create_sheet("산출물")
title(ws_deliv, "산출물")
header(ws_deliv, 3, ["산출물", "Offer A", "Offer B", "Offer C"])
deliv_rows = [
    ("업무 범위", "Y", "Y", "Y"),
    ("필요 입력자료", "Y", "Y", "Y"),
    ("실행 단계", "Y", "Y", "Y"),
    ("AI 보조 단계", "Y", "Y", "Y"),
    ("사람 판단 단계", "Y", "Y", "Y"),
    ("검토 기준", "Y", "Y", "Y"),
    ("금지 업무", "Y", "Y", "Y"),
    ("예외 목록", "Y", "Y", "Y"),
    ("스킬 초안", "Y", "Y", "Y"),
    ("업무 분석", "", "Y", "Y"),
    ("합성 실행", "", "Y", "Y"),
    ("증거 요구사항", "", "Y", "Y"),
    ("누락·충돌 처리", "", "Y", "Y"),
    ("실행자·검토자 역할", "", "Y", "Y"),
    ("수정·재실행", "", "Y", "Y"),
    ("최종 승인 기준", "", "Y", "Y"),
    ("검증된 스킬 카드", "", "Y", "Y"),
    ("운영 가이드", "", "Y", "Y"),
    ("팀 운영 플레이북", "", "", "Y"),
    ("분기 검토 일정", "", "", "Y"),
]
for i, r in enumerate(deliv_rows, start=4):
    for j, v in enumerate(r, start=1):
        ws_deliv.cell(row=i, column=j, value=v)
ws_deliv.column_dimensions["A"].width = 26
for col in "BCD":
    ws_deliv.column_dimensions[col].width = 10

ws_sched = wb.create_sheet("일정")
title(ws_sched, "일정")
header(ws_sched, 3, ["구간", "내용", "기간"])
sched_rows = [
    ("워크숍/분석", "업무 선정·범위·입력 정의", "1~2일 (A)"),
    ("스프린트", "합성 실행·검토·수정·납품", "2~3주 (B)"),
    ("파일럿", "업무 3개·팀 운영", "4~6주 (C)"),
]
for i, r in enumerate(sched_rows, start=4):
    for j, v in enumerate(r, start=1):
        ws_sched.cell(row=i, column=j, value=v)
ws_sched.column_dimensions["A"].width = 16
ws_sched.column_dimensions["B"].width = 40
ws_sched.column_dimensions["C"].width = 20

ws_quote = wb.create_sheet("견적")
title(ws_quote, "견적 (가격 가설 · 계약 아님)")
header(ws_quote, 3, ["항목", "금액(원)", "산식/비고"])
quote_rows = [
    ("공급가액", None, "수동 입력 (가설 범위 내 권장)"),
    ("VAT", "=B4*0.1", "공급가액의 10%"),
    ("총액", "=B4+B5", "공급가액 + VAT"),
    ("착수금", "=B6*0.4", "총액의 40%"),
    ("잔금", "=B6-B7", "총액 - 착수금"),
    ("범위 경고", None, "가설 범위 밖이면 경고 표시"),
]
for i, (a, b, c) in enumerate(quote_rows, start=4):
    ws_quote.cell(row=i, column=1, value=a).font = Font(bold=True)
    if b is not None:
        ws_quote.cell(row=i, column=2, value=b)
    ws_quote.cell(row=i, column=3, value=c)
ws_quote["B4"] = 5000000
ws_quote["B9"] = ("=IF(AND(B4>=VLOOKUP('Offer 선택'!$B$4,$G$4:$I$6,2,FALSE),"
                  "B4<=VLOOKUP('Offer 선택'!$B$4,$G$4:$I$6,3,FALSE)),"
                  "\"가설 범위 내\",\"⚠ 범위 밖 가격 — 가설 범위를 확인하세요\")")
ws_quote["G4"] = "A"
ws_quote["H4"] = 3000000
ws_quote["I4"] = 5000000
ws_quote["G5"] = "B"
ws_quote["H5"] = 5000000
ws_quote["I5"] = 8000000
ws_quote["G6"] = "C"
ws_quote["H6"] = 12000000
ws_quote["I6"] = 20000000
ws_quote.cell(row=9, column=1, value="가격 범위 판정")
ws_quote["B9"].fill = PatternFill("solid", fgColor=YELLOW)
ws_quote.column_dimensions["A"].width = 16
ws_quote.column_dimensions["B"].width = 18
ws_quote.column_dimensions["C"].width = 44
for col in ("G", "H", "I"):
    ws_quote.column_dimensions[col].hidden = True
ws_quote.print_area = "A1:C9"

ws_assume = wb.create_sheet("가정·제외사항")
title(ws_assume, "가정·제외사항")
header(ws_assume, 3, ["항목", "내용"])
assume_rows = [
    ("가정", "합성 데이터 기반 실행, 사람 검토·승인 포함"),
    ("가정", "업무 1개 제한 (A/B), 업무 3개·팀 1개 (C)"),
    ("제외", "backend · 계정 · 인증 · 데이터베이스 · 실시간 AI"),
    ("제외", "파일 업로드 · 기업 연동 · 청구 · 생산 자동화"),
    ("제외", "정확성 보장 · 오류 없음 · 직원 대체"),
    ("주의", "가격은 가설이며 확정 가격이 아닙니다."),
]
for i, (a, b) in enumerate(assume_rows, start=4):
    ws_assume.cell(row=i, column=1, value=a).font = Font(bold=True)
    ws_assume.cell(row=i, column=2, value=b)
ws_assume.column_dimensions["A"].width = 14
ws_assume.column_dimensions["B"].width = 70

wb.save(OUT)
print("saved", OUT, "sheets:", wb.sheetnames)
