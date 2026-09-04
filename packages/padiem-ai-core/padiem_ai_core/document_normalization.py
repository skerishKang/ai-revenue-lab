from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from pathlib import PurePath, PurePosixPath
from typing import Any
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

MAX_DOCUMENT_NAME_CHARS = 120
MAX_DOCUMENT_CHARS = 40_000
MAX_TEXT_DOCUMENT_BYTES = 96 * 1024
MAX_BINARY_DOCUMENT_BYTES = 2 * 1024 * 1024
MAX_PDF_PAGES = 80
MAX_OOXML_ENTRIES = 256
MAX_OOXML_ENTRY_UNCOMPRESSED_BYTES = 1 * 1024 * 1024
MAX_OOXML_TOTAL_UNCOMPRESSED_BYTES = 8 * 1024 * 1024
MAX_XLSX_SHEETS = 20
MAX_XLSX_ROWS = 500
MAX_XLSX_COLUMNS = 50
MAX_XLSX_NONEMPTY_CELLS = 5_000

TEXT_DOCUMENT_MEDIA: dict[str, frozenset[str]] = {
    "text/plain": frozenset({".txt"}),
    "text/markdown": frozenset({".md", ".markdown"}),
    "text/csv": frozenset({".csv"}),
    "application/json": frozenset({".json"}),
}

BINARY_DOCUMENT_MEDIA: dict[str, frozenset[str]] = {
    "application/pdf": frozenset({".pdf"}),
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": frozenset({".docx"}),
    "application/vnd.openxmlformats-officedocument.presentationml.presentation": frozenset({".pptx"}),
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": frozenset({".xlsx"}),
}


class DocumentNormalizationError(ValueError):
    def __init__(self, code: str, safe_message: str) -> None:
        self.code = code
        self.safe_message = safe_message
        super().__init__(safe_message)


@dataclass(frozen=True, slots=True)
class NormalizedDocument:
    name: str
    media_type: str
    text: str = field(repr=False)
    byte_size: int = 0
    source_kind: str = "text"

    def to_public_dict(self) -> dict[str, Any]:
        return {
            "type": "document",
            "name": self.name,
            "media_type": self.media_type,
            "byte_size": self.byte_size,
            "text_chars": len(self.text),
        }


def normalize_document_name(value: Any) -> str:
    if not isinstance(value, str):
        raise DocumentNormalizationError("invalid_name", "Document name must be a string.")
    cleaned = "".join(ch for ch in value.strip() if ch >= " " and ch != "\x7f")
    if not cleaned:
        raise DocumentNormalizationError("name_required", "Document name is required.")
    return cleaned[:MAX_DOCUMENT_NAME_CHARS]


def _validated_media(name: str, media_type: Any, allowed: dict[str, frozenset[str]], *, kind: str) -> str:
    if not isinstance(media_type, str) or media_type not in allowed:
        raise DocumentNormalizationError(f"unsupported_{kind}_media_type", f"Unsupported {kind} document media type.")
    suffix = PurePath(name.lower()).suffix
    if suffix not in allowed[media_type]:
        raise DocumentNormalizationError("media_extension_mismatch", "Document extension does not match media type.")
    return media_type


def validate_document_identity(*, name: Any, media_type: Any, source_kind: str) -> tuple[str, str]:
    safe_name = normalize_document_name(name)
    if source_kind == "text":
        allowed = TEXT_DOCUMENT_MEDIA
    elif source_kind == "binary":
        allowed = BINARY_DOCUMENT_MEDIA
    else:
        raise DocumentNormalizationError("invalid_source_kind", "Document source kind must be text or binary.")
    return safe_name, _validated_media(safe_name, media_type, allowed, kind=source_kind)


def normalize_document_text(value: Any) -> str:
    if not isinstance(value, str):
        raise DocumentNormalizationError("invalid_text", "Document text must be a string.")
    text = value.removeprefix("\ufeff").replace("\r\n", "\n").replace("\r", "\n")
    if not text.strip():
        raise DocumentNormalizationError("empty_document", "Document text is empty.")
    if len(text) > MAX_DOCUMENT_CHARS:
        raise DocumentNormalizationError("text_too_long", "Document text exceeds the character limit.")
    if "\x00" in text:
        raise DocumentNormalizationError("binary_text_rejected", "NUL bytes are not allowed in text documents.")
    bad_controls = sum(1 for ch in text if ord(ch) < 32 and ch not in {"\n", "\t"})
    if bad_controls > max(3, len(text) // 500):
        raise DocumentNormalizationError("excessive_control_characters", "Document contains too many control characters.")
    return text


def normalize_text_document(*, name: Any, media_type: Any, text: Any) -> NormalizedDocument:
    safe_name, safe_media = validate_document_identity(name=name, media_type=media_type, source_kind="text")
    normalized = normalize_document_text(text)
    byte_size = len(normalized.encode("utf-8"))
    if byte_size > MAX_TEXT_DOCUMENT_BYTES:
        raise DocumentNormalizationError("text_bytes_too_large", "Text document exceeds the byte limit.")
    return NormalizedDocument(
        name=safe_name,
        media_type=safe_media,
        text=normalized,
        byte_size=byte_size,
        source_kind="text",
    )


def _append_bounded(parts: list[str], value: str) -> None:
    if not value:
        return
    parts.append(value)
    if sum(len(part) for part in parts) + max(0, len(parts) - 1) > MAX_DOCUMENT_CHARS:
        raise DocumentNormalizationError("extracted_text_too_long", "Extracted document text exceeds the character limit.")


def _safe_ooxml_member(name: str) -> bool:
    if not name or "\\" in name or name.startswith("/") or "//" in name:
        return False
    if len(name) >= 2 and name[1] == ":":
        return False
    parts = PurePosixPath(name).parts
    return all(part not in {"", ".", ".."} for part in parts)


def validate_ooxml_archive(payload: bytes) -> None:
    if not isinstance(payload, (bytes, bytearray)) or not payload or len(payload) > MAX_BINARY_DOCUMENT_BYTES:
        raise DocumentNormalizationError("ooxml_archive_size", "OOXML archive size is out of bounds.")
    try:
        with ZipFile(BytesIO(bytes(payload))) as archive:
            infos = archive.infolist()
            if len(infos) > MAX_OOXML_ENTRIES:
                raise DocumentNormalizationError("ooxml_entry_count", "OOXML archive contains too many entries.")
            total = 0
            for info in infos:
                if not _safe_ooxml_member(info.filename):
                    raise DocumentNormalizationError("ooxml_unsafe_path", "OOXML archive contains an unsafe member path.")
                if info.flag_bits & 0x1:
                    raise DocumentNormalizationError("ooxml_encrypted", "Encrypted OOXML entries are not supported.")
                if info.file_size > MAX_OOXML_ENTRY_UNCOMPRESSED_BYTES:
                    raise DocumentNormalizationError("ooxml_entry_size", "OOXML archive entry exceeds the size limit.")
                total += info.file_size
                if total > MAX_OOXML_TOTAL_UNCOMPRESSED_BYTES:
                    raise DocumentNormalizationError("ooxml_total_size", "OOXML archive exceeds the total uncompressed size limit.")
                lowered = info.filename.lower()
                if lowered.endswith(".xml") or lowered.endswith(".rels"):
                    xml = archive.read(info)
                    if b"<!doctype" in xml.lower():
                        raise DocumentNormalizationError("ooxml_dtd_rejected", "DTDs are not supported in OOXML documents.")
    except DocumentNormalizationError:
        raise
    except (BadZipFile, OSError, ValueError, RuntimeError) as exc:
        raise DocumentNormalizationError("ooxml_malformed", "Malformed OOXML ZIP archive.") from exc


def _parse_xml(xml: bytes) -> ElementTree.Element:
    if b"<!doctype" in xml.lower():
        raise DocumentNormalizationError("ooxml_dtd_rejected", "DTDs are not supported in OOXML documents.")
    try:
        return ElementTree.fromstring(xml)
    except ElementTree.ParseError as exc:
        raise DocumentNormalizationError("ooxml_invalid_xml", "Invalid OOXML XML part.") from exc


def _local_name(element: ElementTree.Element) -> str:
    return element.tag.rsplit("}", 1)[-1]


def extract_docx_text(payload: bytes) -> str:
    validate_ooxml_archive(payload)
    try:
        with ZipFile(BytesIO(payload)) as archive:
            try:
                root = _parse_xml(archive.read("word/document.xml"))
            except KeyError as exc:
                raise DocumentNormalizationError("docx_missing_part", "DOCX is missing word/document.xml.") from exc
    except DocumentNormalizationError:
        raise
    except (BadZipFile, OSError, ValueError, RuntimeError) as exc:
        raise DocumentNormalizationError("ooxml_malformed", "Malformed OOXML ZIP archive.") from exc

    paragraphs: list[str] = []
    for paragraph in root.iter():
        if _local_name(paragraph) != "p":
            continue
        pieces = [node.text or "" for node in paragraph.iter() if _local_name(node) == "t"]
        value = "".join(pieces)
        if value:
            _append_bounded(paragraphs, value)
    text = "\n".join(paragraphs).strip()
    if not text:
        raise DocumentNormalizationError("docx_empty", "DOCX contains no readable text.")
    return text


def _slide_sort_key(name: str) -> tuple[int, str]:
    stem = PurePosixPath(name).stem
    suffix = stem.removeprefix("slide")
    return (int(suffix) if suffix.isdigit() else 10**9, name)


def extract_pptx_text(payload: bytes) -> str:
    validate_ooxml_archive(payload)
    try:
        with ZipFile(BytesIO(payload)) as archive:
            slide_names = sorted(
                (
                    name
                    for name in archive.namelist()
                    if name.startswith("ppt/slides/slide") and name.endswith(".xml") and "/_rels/" not in name
                ),
                key=_slide_sort_key,
            )
            if not slide_names:
                raise DocumentNormalizationError("pptx_missing_slides", "PPTX contains no slide XML parts.")
            slides: list[str] = []
            for slide_name in slide_names:
                root = _parse_xml(archive.read(slide_name))
                pieces = [node.text or "" for node in root.iter() if _local_name(node) == "t"]
                slide = "\n".join(piece for piece in pieces if piece).strip()
                if slide:
                    _append_bounded(slides, slide)
    except DocumentNormalizationError:
        raise
    except (BadZipFile, OSError, ValueError, RuntimeError) as exc:
        raise DocumentNormalizationError("ooxml_malformed", "Malformed OOXML ZIP archive.") from exc
    text = "\n".join(slides).strip()
    if not text:
        raise DocumentNormalizationError("pptx_empty", "PPTX contains no readable text.")
    return text


def _extract_pdf_text(payload: bytes) -> str:
    if not payload.startswith(b"%PDF-"):
        raise DocumentNormalizationError("pdf_magic_mismatch", "PDF magic does not match the declared media type.")
    try:
        from pypdf import PdfReader
    except ModuleNotFoundError as exc:
        raise DocumentNormalizationError("document_dependency_unavailable", "PDF extraction dependency is unavailable.") from exc
    try:
        reader = PdfReader(BytesIO(payload), strict=True)
        if reader.is_encrypted:
            raise DocumentNormalizationError("pdf_encrypted", "Encrypted PDF documents are not supported.")
        page_count = len(reader.pages)
        if page_count < 1:
            raise DocumentNormalizationError("pdf_no_pages", "PDF contains no readable pages.")
        if page_count > MAX_PDF_PAGES:
            raise DocumentNormalizationError("pdf_page_limit", "PDF exceeds the page limit.")
        parts: list[str] = []
        for page in reader.pages:
            extracted = (page.extract_text() or "").strip()
            if extracted:
                _append_bounded(parts, extracted)
        text = "\n\n".join(parts).strip()
        if not text:
            raise DocumentNormalizationError("pdf_empty_text", "PDF contains no extractable text; OCR is not enabled.")
        return text
    except DocumentNormalizationError:
        raise
    except Exception as exc:
        raise DocumentNormalizationError("pdf_invalid", "PDF could not be parsed safely.") from exc


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def _extract_xlsx_text(payload: bytes) -> str:
    validate_ooxml_archive(payload)
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise DocumentNormalizationError("document_dependency_unavailable", "XLSX extraction dependency is unavailable.") from exc
    try:
        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True, keep_links=False)
    except Exception as exc:
        raise DocumentNormalizationError("xlsx_invalid", "XLSX could not be parsed safely.") from exc
    try:
        if len(workbook.sheetnames) < 1:
            raise DocumentNormalizationError("xlsx_no_sheets", "XLSX contains no worksheets.")
        if len(workbook.sheetnames) > MAX_XLSX_SHEETS:
            raise DocumentNormalizationError("xlsx_sheet_limit", "XLSX exceeds the worksheet limit.")
        parts: list[str] = []
        nonempty_cells = 0
        for worksheet in workbook.worksheets:
            max_row = int(worksheet.max_row or 0)
            max_column = int(worksheet.max_column or 0)
            if max_row > MAX_XLSX_ROWS or max_column > MAX_XLSX_COLUMNS:
                raise DocumentNormalizationError("xlsx_dimension_limit", "XLSX worksheet dimensions exceed the allowed range.")
            for row in worksheet.iter_rows(max_row=min(max_row, MAX_XLSX_ROWS), max_col=min(max_column, MAX_XLSX_COLUMNS)):
                for cell in row:
                    value = _cell_text(cell.value)
                    if not value:
                        continue
                    nonempty_cells += 1
                    if nonempty_cells > MAX_XLSX_NONEMPTY_CELLS:
                        raise DocumentNormalizationError("xlsx_cell_limit", "XLSX contains too many non-empty cells.")
                    _append_bounded(parts, f"[{worksheet.title}!{cell.coordinate}] {value}")
        text = "\n".join(parts).strip()
        if not text:
            raise DocumentNormalizationError("xlsx_empty", "XLSX contains no readable values.")
        return text
    finally:
        workbook.close()


def extract_binary_document(*, name: Any, media_type: Any, payload: Any) -> NormalizedDocument:
    safe_name, safe_media = validate_document_identity(name=name, media_type=media_type, source_kind="binary")
    if not isinstance(payload, (bytes, bytearray)):
        raise DocumentNormalizationError("invalid_binary_payload", "Binary document payload must be bytes.")
    binary = bytes(payload)
    if not binary:
        raise DocumentNormalizationError("empty_document", "Document payload is empty.")
    if len(binary) > MAX_BINARY_DOCUMENT_BYTES:
        raise DocumentNormalizationError("binary_too_large", "Binary document exceeds the byte limit.")

    if safe_media == "application/pdf":
        text = _extract_pdf_text(binary)
    elif safe_media == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        text = extract_docx_text(binary)
    elif safe_media == "application/vnd.openxmlformats-officedocument.presentationml.presentation":
        text = extract_pptx_text(binary)
    elif safe_media == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        text = _extract_xlsx_text(binary)
    else:  # pragma: no cover - guarded by media validation
        raise DocumentNormalizationError("unsupported_binary_media_type", "Unsupported binary document media type.")

    normalized = normalize_document_text(text)
    return NormalizedDocument(
        name=safe_name,
        media_type=safe_media,
        text=normalized,
        byte_size=len(binary),
        source_kind="binary",
    )
